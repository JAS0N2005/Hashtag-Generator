import re
import csv
import os
import pandas as pd
from openpyxl import load_workbook
from config import FINAL_MASTER_SHEET_NAME, CONFLICT_HASHTAG_MAP, EXCLUDED_TYPE_HASHTAGS

def sanitize_filename(name: str) -> str:
    """
    Strip only the chars < > : \" / \\ | ? * from `name`,
    then trim any trailing spaces or dots.
    """
    # remove illegal Windows filename chars
    name = re.sub(r'[<>:"/\\|?*]', '', name)
    # remove trailing spaces or periods
    return name.rstrip(' .')

def count_rows_xlsx(path: str, candidates: list[str]) -> int:
    """
    Return number of data rows (excluding header) in the first sheet
    whose name (stripped & lowercased) appears in candidates.
    Scans every worksheet (visible or hidden) and normalizes titles.
    """
    wb = load_workbook(path, read_only=True, data_only=True)

    # Normalize your candidate names (strip spaces, lowercase)
    norm_cands = {c.strip().lower() for c in candidates}

    # Iterate through every worksheet, not just wb.sheetnames
    for ws in wb.worksheets:
        title_norm = ws.title.strip().lower()
        try:
            if title_norm in norm_cands:
                # max_row might be None for an empty sheet
                max_row = ws.max_row or 0
                # subtract 1 for header, but never go below zero
                return max_row - 1
        except:
            # If we never found a match, treat as “missing” (you could also raise)
            return 0


def init_logger(log_path):
    """Create parent folder, write header if file doesn’t exist, and return an open file + writer."""
    os.makedirs(os.path.dirname(log_path), exist_ok=True)
    file_existed = os.path.exists(log_path)
    f = open(log_path, mode='a', newline='', encoding='utf-8')
    writer = csv.writer(f)
    if not file_existed:
        writer.writerow(["City", "MasterCount", "OtherCount","Tally Status"])
    return f, writer

# ─── Sanitizer: strip everything except letters, digits, underscore and “#” ───
_SANITIZE_RE = re.compile(r'[^\w#]', flags=re.UNICODE)

def sanitize_hashtag(tag: str) -> str:
    """
    Clean a hashtag string by removing all special characters
    except underscore (_) and the leading hash (#).
    """
    return _SANITIZE_RE.sub('', tag)

# ─── Precompute lowercase maps for faster lookups ──────────────────────────────
CONFLICT_HASHTAG_MAP = {
    dominant.lower(): {u.lower() for u in unwanted}
    for dominant, unwanted in CONFLICT_HASHTAG_MAP.items()
}
_exclude_lower = {h.lower() for h in EXCLUDED_TYPE_HASHTAGS}

def extract_food_sheets(sheets, keywords):
    """
    Return list of DataFrames for sheets whose names contain any keyword,
    and list of sheet names to remove.
    """
    matches = []
    to_remove = []
    for name, df in sheets.items():
        name_clean = name.strip().lower()
        for kw in keywords:
            if kw.strip().lower() in name_clean:
                matches.append(normalize_columns(df))
                to_remove.append(name)
                break
    return matches, to_remove

def remove_sheets(sheets, to_remove):
    """Remove specified sheet names from a sheets dict."""
    for name in to_remove:
        sheets.pop(name, None)

def consolidate_processed_sheets(sheets: dict) -> dict:
    """
    Combine all DataFrames into one final master sheet.
    """
    combined = pd.concat(sheets.values(), ignore_index=True) if sheets else pd.DataFrame()
    return {FINAL_MASTER_SHEET_NAME: combined}

def normalize_columns(df):
    """Standardize column names by stripping whitespace."""
    df.columns = [col.strip() for col in df.columns]
    return df

def extract_existing(cell):
    """Return a list and set of existing hashtags from a cell."""
    if not isinstance(cell, str):
        return [], set()
    tags = [line.strip() for line in cell.splitlines() if line.strip().startswith('#')]
    return tags, set(tags)

def deduplicate_hashtags(hashtags, exclude_hashtags=None, conflict_map=None):
    """
    0) Sanitize all tags (remove special chars except underscore)
    1) Remove conflicts based on CONFLICT_HASHTAG_MAP
    2) Auto-prefer plurals: drop any tag if its plural (s or es) also exists
    3) Exclude any in EXCLUDED_TYPE_HASHTAGS
    4) Remove remaining exact duplicates
    """
    # 0) Sanitize
    hashtags = [sanitize_hashtag(h) for h in hashtags]

    # 1) Conflict removal
    lowers = {h.lower() for h in hashtags}
    filtered = []
    for h in hashtags:
        h_low = h.lower()
        drop = False
        for dominant_lower, unwanted_lower in CONFLICT_HASHTAG_MAP.items():
            if dominant_lower in lowers and h_low in unwanted_lower:
                drop = True
                break
        if not drop:
            filtered.append(h)
    hashtags = filtered

    # 2) Auto-prefer plurals: drop singular if plural exists (s or es)
    lowers = {h.lower() for h in hashtags}
    cleaned = []
    for h in hashtags:
        h_low = h.lower()
        # if a plural version exists, skip this tag
        if (h_low + 's') in lowers or (h_low + 'es') in lowers:
            continue
        cleaned.append(h)
    hashtags = cleaned

    # 3) Exclude explicit tags
    hashtags = [h for h in hashtags if h.lower() not in _exclude_lower]

    # 4) Final dedupe preserving order
    seen = set()
    result = []
    for h in hashtags:
        key = h.lower()
        if key not in seen:
            seen.add(key)
            result.append(h)

    return result

def move_column(df, col_name, new_index):
    """
    Move a column in a DataFrame to a new index, preserving order of others.
    """
    cols = list(df.columns)
    if col_name not in cols:
        return df
    cols.remove(col_name)
    new_index = max(0, min(new_index, len(cols)))
    cols.insert(new_index, col_name)
    return df[cols]
